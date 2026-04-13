"""
Advanced Layout-Aware Document Parser.

Provides frontier-quality layout analysis, structure extraction,
and metadata enrichment for PDF and Office documents.
Inspired by RAGFlow DeepDoc architecture.
"""

import os
import re
import json
import logging
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple
from enum import Enum

logger = logging.getLogger(__name__)


class LayoutType(str, Enum):
    TITLE = "title"
    HEADING = "heading"
    TEXT = "text"
    PARAGRAPH = "paragraph"
    TABLE = "table"
    TABLE_CAPTION = "table_caption"
    FIGURE = "figure"
    FIGURE_CAPTION = "figure_caption"
    HEADER = "header"
    FOOTER = "footer"
    REFERENCE = "reference"
    EQUATION = "equation"
    LIST = "list"
    LIST_ITEM = "list_item"
    BLOCK_QUOTE = "block_quote"
    CODE = "code"
    PAGE_BREAK = "page_break"
    UNKNOWN = "unknown"


@dataclass
class LayoutBlock:
    """A detected layout block with type and position."""

    layout_type: LayoutType
    text: str
    page: int = 0
    bbox: Optional[Tuple[int, int, int, int]] = None
    confidence: float = 1.0
    children: List["LayoutBlock"] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class LayoutAnalysisResult:
    """Result of layout analysis."""

    blocks: List[LayoutBlock]
    metadata: Dict[str, Any] = field(default_factory=dict)
    page_count: int = 0
    table_count: int = 0
    figure_count: int = 0
    heading_tree: Dict[str, Any] = field(default_factory=dict)


@dataclass
class StructureAnalysisResult:
    """Document structure analysis result."""

    sections: List[Dict[str, Any]] = field(default_factory=list)
    table_of_contents: List[Dict[str, Any]] = field(default_factory=list)
    hierarchy: Dict[str, Any] = field(default_factory=dict)
    reading_order: List[str] = field(default_factory=list)


class PDFLayoutAnalyzer:
    """PDF layout analysis using available backends."""

    def __init__(self):
        self._pymupdf_available = False
        self._pymupdf = None

        try:
            import fitz

            self._pymupdf = fitz
            self._pymupdf_available = True
        except ImportError:
            pass

    @property
    def available(self) -> bool:
        return self._pymupdf_available

    async def analyze(
        self,
        content: bytes,
    ) -> LayoutAnalysisResult:
        if not self._pymupdf_available:
            return LayoutAnalysisResult(blocks=[])

        try:
            with self._pymupdf.open(stream=content, filetype="pdf") as doc:
                blocks = []
                metadata = {}
                page_count = len(doc)
                table_count = 0
                figure_count = 0

                for page_num, page in enumerate(doc):
                    text_dict = page.get_text("dict")

                    for block in text_dict.get("blocks", []):
                        block_type = block.get("type", 0)

                        if block_type == 0:
                            bbox = block.get("bbox", (0, 0, 0, 0))
                            lines = block.get("lines", [])

                            text_parts = []
                            for line in lines:
                                for span in line.get("spans", []):
                                    text_parts.append(span.get("text", ""))

                            text = "".join(text_parts)

                            layout_type = self._classify_text_block(
                                text, bbox, page_num + 1
                            )

                            block_obj = LayoutBlock(
                                layout_type=layout_type,
                                text=text.strip(),
                                page=page_num + 1,
                                bbox=tuple(bbox) if bbox else None,
                            )

                            if layout_type == LayoutType.TABLE:
                                table_count += 1
                            elif layout_type == LayoutType.FIGURE:
                                figure_count += 1

                            blocks.append(block_obj)

                heading_tree = self._build_heading_tree(blocks)

                return LayoutAnalysisResult(
                    blocks=blocks,
                    metadata=metadata,
                    page_count=page_count,
                    table_count=table_count,
                    figure_count=figure_count,
                    heading_tree=heading_tree,
                )

        except Exception as e:
            logger.error("Failed to analyze PDF layout: %s", e)
            return LayoutAnalysisResult(blocks=[])

    def _classify_text_block(self, text: str, bbox: Tuple, page: int) -> LayoutType:
        text = text.strip()
        if not text:
            return LayoutType.UNKNOWN

        if len(text) < 100 and (
            text.isupper() or text.startswith("#") or re.match(r"^[A-Z][^.!?]*$", text)
        ):
            return LayoutType.TITLE

        if re.match(r"^#{1,6}\s+", text):
            return LayoutType.HEADING

        if re.match(r"^\d+\.\s+", text):
            return LayoutType.LIST_ITEM

        if "|" in text and text.count("|") > 2:
            return LayoutType.TABLE

        if text.startswith("[") and text.endswith("]"):
            return LayoutType.REFERENCE

        return LayoutType.PARAGRAPH

    def _build_heading_tree(self, blocks: List[LayoutBlock]) -> Dict[str, Any]:
        tree = {"root": {"children": [], "level": 0}}
        current_path = []

        for block in blocks:
            if block.layout_type == LayoutType.HEADING:
                first_word = block.text.split()[0] if block.text else ""
                level = len(first_word) if first_word else 1
                if level not in [1, 2, 3, 4, 5, 6]:
                    level = 1

                node = {
                    "text": block.text,
                    "level": level,
                    "page": block.page,
                    "children": [],
                }

                while current_path and current_path[-1]["level"] >= level:
                    current_path.pop()

                if current_path:
                    current_path[-1]["children"].append(node)
                else:
                    tree["root"]["children"].append(node)

                current_path.append(node)

        return tree


class DOCXStructureAnalyzer:
    """DOCX structure analysis for headings, tables, styles."""

    def __init__(self):
        pass

    async def analyze(
        self,
        content: bytes,
    ) -> StructureAnalysisResult:
        sections = []
        table_of_contents = []
        hierarchy = {"root": {"children": []}}
        reading_order = []

        try:
            with zipfile.ZipFile(BytesIO(content), "r") as docx:
                with docx.open("word/document.xml") as doc_xml:
                    tree = ET.parse(doc_xml)
                    root = tree.getroot()

                    ns = {
                        "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
                    }

                    headings = root.findall(".//w:pStyle[@w:val='Heading']/..", ns)

                    for para in root.findall(".//w:p", ns):
                        p_pr = para.find("w:pPr", ns)

                        style = None
                        if p_pr is not None:
                            p_style = p_pr.find("w:pStyle", ns)
                            if p_style is not None:
                                style = p_style.get(
                                    "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val"
                                )

                        text_parts = []
                        for text in para.findall(".//w:t", ns):
                            if text.text:
                                text_parts.append(text.text)

                        text = "".join(text_parts)

                        if style and "heading" in style.lower():
                            level = int(re.search(r"\d+", style).group() or "1")
                            sections.append(
                                {
                                    "text": text,
                                    "level": level,
                                    "type": "heading",
                                }
                            )
                            table_of_contents.append({"text": text, "level": level})
                        elif text.strip():
                            sections.append(
                                {
                                    "text": text,
                                    "level": 0,
                                    "type": "paragraph",
                                }
                            )

                        reading_order.append(text)

                tables = self._extract_tables(docx)
                for table in tables:
                    sections.append(
                        {
                            "type": "table",
                            "rows": table,
                        }
                    )

        except Exception as e:
            logger.error("Failed to analyze DOCX structure: %s", e)

        return StructureAnalysisResult(
            sections=sections,
            table_of_contents=table_of_contents,
            hierarchy=hierarchy,
            reading_order=reading_order,
        )

    def _extract_tables(self, docx) -> List[List[str]]:
        tables = []

        try:
            for name in docx.namelist():
                if "word/document.xml" in name:
                    continue

                if name.startswith("word/") and name.endswith(".xml"):
                    with docx.open(name) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()

                        ns = {
                            "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
                        }

                        for tbl in root.findall(".//w:tbl", ns):
                            rows = []
                            for tr in tbl.findall("w:tr", ns):
                                cells = []
                                for tc in tr.findall("w:tc", ns):
                                    cell_text = "".join(
                                        p.text or ""
                                        for p in tc.findall(".//w:p/w:t", ns)
                                    )
                                    cells.append(cell_text)
                                rows.append(cells)

                            if rows:
                                tables.append(rows)

        except Exception as e:
            logger.error("Failed to extract tables from DOCX: %s", e)

        return tables


class XLSXStructureAnalyzer:
    """XLSX structure analysis for sheets and ranges."""

    def __init__(self):
        pass

    async def analyze(
        self,
        content: bytes,
    ) -> StructureAnalysisResult:
        sheets = []
        table_of_contents = []
        reading_order = []

        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(content))

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                sheet_info = {
                    "name": sheet_name,
                    "type": "sheet",
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                }
                sheets.append(sheet_info)
                table_of_contents.append(sheet_info)

                has_data = False
                for row in ws.iter_rows(max_row=min(10, ws.max_row)):
                    row_data = [cell.value for cell in row]
                    if any(row_data):
                        has_data = True
                        reading_order.append(row_data)

        except Exception:
            pass

        return StructureAnalysisResult(
            sections=sheets,
            table_of_contents=table_of_contents,
            hierarchy={"sheets": sheets},
            reading_order=reading_order,
        )


class PPTXStructureAnalyzer:
    """PPTX structure analysis for slides and notes."""

    def __init__(self):
        pass

    async def analyze(
        self,
        content: bytes,
    ) -> StructureAnalysisResult:
        sections = []
        table_of_contents = []
        hierarchy = {"slides": []}
        reading_order = []

        try:
            with zipfile.ZipFile(BytesIO(content), "r") as pptx:
                slide_files = sorted(
                    [f for f in pptx.namelist() if f.startswith("ppt/slides/slide")],
                    key=lambda x: int(
                        re.search(r"slide(\d+)", x).group(1)
                        if re.search(r"slide(\d+)", x)
                        else 0
                    ),
                )

                ns = {
                    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
                }

                for slide_path in slide_files:
                    with pptx.open(slide_path) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()

                        slide_num = int(re.search(r"slide(\d+)", slide_path).group(1))

                        shapes = []
                        for shape in root.findall(".//p:sp", ns):
                            text_elems = shape.findall(".//a:p/a:r/a:t", ns)
                            text = "".join(t.text or "" for t in text_elems)
                            if text.strip():
                                shapes.append(text.strip())

                        slide_info = {
                            "slide_number": slide_num,
                            "shapes": shapes,
                        }
                        sections.append(slide_info)
                        table_of_contents.append(slide_info)
                        hierarchy["slides"].append(slide_info)

                        reading_order.extend(shapes)

                notes_files = [f for f in pptx.namelist() if "notesSlide" in f]

                for notes_path in notes_files:
                    with pptx.open(notes_path) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()

                        text_elems = root.findall(".//a:t", ns)
                        notes = "".join(t.text or "" for t in text_elems)

                        if notes.strip():
                            sections.append(
                                {
                                    "type": "notes",
                                    "content": notes.strip(),
                                }
                            )

        except Exception as e:
            logger.error("Failed to analyze PPTX structure: %s", e)

        return StructureAnalysisResult(
            sections=sections,
            table_of_contents=table_of_contents,
            hierarchy=hierarchy,
            reading_order=reading_order,
        )


from io import BytesIO


class UnifiedLayoutParser:
    """Unified parser coordinating all layout analyzers."""

    def __init__(self):
        self.pdf_analyzer = PDFLayoutAnalyzer()
        self.docx_analyzer = DOCXStructureAnalyzer()
        self.xlsx_analyzer = XLSXStructureAnalyzer()
        self.pptx_analyzer = PPTXStructureAnalyzer()

    def get_layout_analysis(
        self, content: bytes, filename: str
    ) -> Optional[LayoutAnalysisResult]:
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".pdf":
            return asyncio.get_event_loop().run_until_complete(
                self.pdf_analyzer.analyze(content)
            )

        return None

    def get_structure_analysis(
        self, content: bytes, filename: str
    ) -> Optional[StructureAnalysisResult]:
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".docx":
            return asyncio.get_event_loop().run_until_complete(
                self.docx_analyzer.analyze(content)
            )
        elif ext == ".xlsx":
            return asyncio.get_event_loop().run_until_complete(
                self.xlsx_analyzer.analyze(content)
            )
        elif ext == ".pptx":
            return asyncio.get_event_loop().run_until_complete(
                self.pptx_analyzer.analyze(content)
            )

        return None


import asyncio


def get_layout_parser() -> UnifiedLayoutParser:
    return UnifiedLayoutParser()


def get_pdf_layout_analyzer() -> PDFLayoutAnalyzer:
    return PDFLayoutAnalyzer()


def get_docx_structure_analyzer() -> DOCXStructureAnalyzer:
    return DOCXStructureAnalyzer()


def get_xlsx_structure_analyzer() -> XLSXStructureAnalyzer:
    return XLSXStructureAnalyzer()


def get_pptx_structure_analyzer() -> PPTXStructureAnalyzer:
    return PPTXStructureAnalyzer()
