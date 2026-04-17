"""
Document Parser - Multi-format document parsing with LlamaIndex + Docling.

Uses:
- LlamaIndex readers for document parsing
- Docling for advanced PDF/OCR
- Custom fallback parsers
"""

import logging
import os
import re
import zipfile
import tempfile
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# LlamaIndex readers (mandatory dependency)
# Use SimpleDirectoryReader which auto-detects file types
from llama_index.core import SimpleDirectoryReader
from llama_index.core.readers import StringIterableReader

# Docling imports
try:
    from docling.document import PdfConverter
    from docling.datamodel.base import PdfNote

    DOCLING_AVAILABLE = True
except ImportError:
    PdfConverter = None
    PdfNote = None
    DOCLING_AVAILABLE = False

# LlamaIndex is mandatory
LLAMA_INDEX_AVAILABLE = True


@dataclass
class ParsedDocumentResult:
    """Unified parsing result."""

    text: str
    metadata: Dict[str, Any] = field(default_factory=dict)
    tables: List[List[str]] = field(default_factory=list)
    images: List[Dict[str, Any]] = field(default_factory=list)
    used_llama_index: bool = False
    used_docling: bool = False
    error: Optional[str] = None
    links: List[str] = field(default_factory=list)
    layout_analysis: Optional[Dict[str, Any]] = field(default_factory=dict)
    structure_analysis: Optional[Dict[str, Any]] = field(default_factory=dict)


class LlamaIndexParser:
    """LlamaIndex-based document parser."""

    def __init__(self):
        self._readers: Dict[str, Any] = {}

    @property
    def available(self) -> bool:
        return LLAMA_INDEX_AVAILABLE

    def _get_reader(self, ext: str):
        if ext in self._readers:
            return self._readers[ext]

        if not self.available:
            return None

        reader_map = {
            ".md": MarkdownReader,
            ".markdown": MarkdownReader,
            ".pdf": PDFReader,
            ".docx": DocxReader,
            ".pptx": PptxReader,
            ".csv": CSVReader,
            ".html": HTMLTagReader,
            ".htm": HTMLTagReader,
            ".txt": FlatReader,
            ".text": FlatReader,
        }

        reader_cls = reader_map.get(ext)
        if reader_cls:
            try:
                self._readers[ext] = reader_cls()
            except Exception as e:
                logger.error("Failed to initialize LlamaIndex reader for %s: %s", ext, e)

        return self._readers.get(ext)

    async def parse(
        self,
        content: bytes,
        filename: str,
    ) -> ParsedDocumentResult:
        if not self.available:
            return ParsedDocumentResult(text="", error="LlamaIndex not available")

        ext = os.path.splitext(filename)[1].lower()
        reader = self._get_reader(ext)

        if not reader:
            return ParsedDocumentResult(text="", error=f"No reader for {ext}")

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            try:
                docs = reader.load_data(tmp_path)
                text = "\n\n".join(doc.text for doc in docs)
                return ParsedDocumentResult(
                    text=text,
                    metadata={"doc_count": len(docs)},
                    used_llama_index=True,
                )
            finally:
                os.unlink(tmp_path)
        except Exception as e:
            return ParsedDocumentResult(text="", error=str(e))


class DoclingParser:
    """Docling-based advanced document parser with optional OCR."""

    def __init__(self, use_ocr: bool = False):
        self._converter = None
        self._use_ocr = use_ocr

    @property
    def available(self) -> bool:
        return DOCLING_AVAILABLE

    def _get_converter(self):
        if not self.available:
            return None

        if self._converter is None:
            try:
                if self._use_ocr:
                    self._converter = PdfConverter(
                        ocr=DoclingProxyOcr(),
                    )
                else:
                    self._converter = PdfConverter()
            except Exception as e:
                logger.error("Failed to initialize Docling PdfConverter: %s", e)

        return self._converter

    async def parse(
        self,
        content: bytes,
    ) -> ParsedDocumentResult:
        converter = self._get_converter()

        if not converter:
            return ParsedDocumentResult(text="", error="Docling not available")

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            try:
                note = PdfNote.from_path(tmp_path)
                result = converter.convert(note)
                text = result.document.export_to_markdown()
                return ParsedDocumentResult(
                    text=text,
                    metadata={"page_count": len(result.pages)},
                    used_docling=True,
                )
            finally:
                os.unlink(tmp_path)
        except Exception as e:
            return ParsedDocumentResult(text="", error=str(e))

    async def parse_with_elements(
        self,
        content: bytes,
    ) -> tuple[str, list[dict], dict]:
        """Parse PDF and return (text, elements, metadata).

        Returns raw tuples for callers to wrap in their own result types.
        """
        converter = self._get_converter()

        if not converter:
            return "", [], {}

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            try:
                note = PdfNote.from_path(tmp_path)
                result = converter.convert(note)
                text = result.document.export_to_markdown()

                elements = []
                for item in result.document.iter_inferred_terms():
                    elements.append(
                        {
                            "element_id": item.id or "",
                            "type": item.label or "unknown",
                            "label": item.text,
                            "confidence": getattr(item, "score", 0.0),
                        }
                    )

                metadata = {"page_count": len(result.pages)}
                return text, elements, metadata
            finally:
                os.unlink(tmp_path)
        except Exception as e:
            return "", [], {"error": str(e)}


class FallbackParser:
    """Custom fallback parsers."""

    def parse(
        self,
        content: bytes,
        filename: str,
    ) -> ParsedDocumentResult:
        ext = os.path.splitext(filename)[1].lower()

        parsers = {
            ".pdf": self._parse_pdf,
            ".docx": self._parse_docx,
            ".xlsx": self._parse_xlsx,
            ".xlsm": self._parse_xlsx,
            ".pptx": self._parse_pptx,
            ".md": self._parse_markdown,
            ".markdown": self._parse_markdown,
            ".html": self._parse_html,
            ".htm": self._parse_html,
            ".json": self._parse_json,
            ".yaml": self._parse_yaml,
            ".yml": self._parse_yaml,
            ".xml": self._parse_xml,
        }

        parser = parsers.get(ext, self._parse_text)
        return parser(content)

    def _parse_text(self, content: bytes) -> ParsedDocumentResult:
        text = content.decode("utf-8", errors="replace")
        return ParsedDocumentResult(text=text)

    def _parse_pdf(self, content: bytes) -> ParsedDocumentResult:
        try:
            import pdfplumber

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            text_parts = []
            try:
                with pdfplumber.open(tmp_path) as doc:
                    for page in doc.pages:
                        text_parts.append(page.extract_text() or "")
            finally:
                os.unlink(tmp_path)

            return ParsedDocumentResult(text="\n".join(text_parts))
        except Exception as e:
            return ParsedDocumentResult(text="", error=str(e))

    def _parse_docx(self, content: bytes) -> ParsedDocumentResult:
        try:
            import zipfile
            import xml.etree.ElementTree as ET

            with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            paragraphs = []
            tables = []
            current_table = []

            try:
                with zipfile.ZipFile(tmp_path) as z:
                    if "word/document.xml" in z.namelist():
                        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

                        with z.open("word/document.xml") as f:
                            tree = ET.parse(f)

                        for elem in tree.iter():
                            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag

                            if tag == "p":
                                text = "".join(
                                    t.text or ""
                                    for t in elem.iter()
                                    if t.tag.endswith("}t") and t.text
                                )
                                if text.strip():
                                    style = elem.attrib.get(
                                        "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}pStyle",
                                        "",
                                    )
                                    if (
                                        "Heading" in style
                                        or style.startswith("1")
                                        or style.startswith("Title")
                                    ):
                                        paragraphs.append(f"## {text.strip()}")
                                    else:
                                        paragraphs.append(text.strip())

                            elif tag == "tbl":
                                table_rows = []
                                for row in elem.findall(".//w:tr", ns):
                                    cells = []
                                    for cell in row.findall(".//w:tc", ns):
                                        cell_text = "".join(
                                            t.text or ""
                                            for t in cell.iter()
                                            if t.tag.endswith("}t") and t.text
                                        )
                                        cells.append(cell_text.strip())
                                    if cells:
                                        table_rows.append(cells)
                                if table_rows:
                                    tables.append(table_rows)

                            elif tag == "pPr":
                                pass

            finally:
                os.unlink(tmp_path)

            output_parts = []

            for p in paragraphs:
                output_parts.append(p)

            if tables:
                output_parts.append("")
                for table in tables:
                    for row in table:
                        output_parts.append("| " + " | ".join(row) + " |")
                    output_parts.append("")

            return ParsedDocumentResult(
                text="\n\n".join(output_parts),
                metadata={"paragraphs": len(paragraphs), "tables": len(tables)},
            )
        except Exception as e:
            return ParsedDocumentResult(text="", error=str(e))

    def _parse_xlsx(self, content: bytes) -> ParsedDocumentResult:
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            output_parts = []
            total_sheets = 0
            total_rows = 0

            try:
                wb = openpyxl.load_workbook(tmp_path, data_only=True)
                total_sheets = len(wb.sheetnames)

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]

                    if total_sheets > 1:
                        output_parts.append(f"## Sheet: {sheet_name}")
                        output_parts.append("")

                    max_row = sheet.max_row
                    max_col = sheet.max_column

                    if max_row == 0:
                        continue

                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        if any(cell is not None for cell in row):
                            formatted_row = []
                            for cell in row:
                                if cell is None:
                                    formatted_row.append("")
                                elif isinstance(cell, (int, float)):
                                    formatted_row.append(str(cell))
                                else:
                                    formatted_row.append(str(cell)[:200])

                            if row_idx == 1:
                                output_parts.append(
                                    "| " + " | ".join(["---"] * len(formatted_row)) + " |"
                                )

                            output_parts.append("| " + " | ".join(formatted_row) + " |")
                            total_rows += 1

                    if (
                        total_sheets > 1
                        and wb.sheetnames.index(sheet_name) < len(wb.sheetnames) - 1
                    ):
                        output_parts.append("")

            finally:
                os.unlink(tmp_path)

            return ParsedDocumentResult(
                text="\n".join(output_parts),
                metadata={"sheets": total_sheets, "total_rows": total_rows},
            )
        except Exception as e:
            return ParsedDocumentResult(text="", error=str(e))

    def _parse_pptx(self, content: bytes) -> ParsedDocumentResult:
        try:
            import zipfile
            import xml.etree.ElementTree as ET

            with tempfile.NamedTemporaryFile(delete=False, suffix=".pptx") as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            slides = []
            notes = []
            layouts = []

            try:
                with zipfile.ZipFile(tmp_path) as z:
                    slide_files = sorted(
                        [
                            n
                            for n in z.namelist()
                            if n.startswith("ppt/slides/slide") and n.endswith(".xml")
                        ]
                    )

                    for slide_num, slide_file in enumerate(slide_files, 1):
                        with z.open(slide_file) as f:
                            tree = ET.parse(f)

                        texts = []
                        for elem in tree.iter():
                            if elem.tag.endswith("}t") and elem.text and elem.text.strip():
                                texts.append(elem.text.strip())

                        if texts:
                            slides.append(f"### Slide {slide_num}\n" + "\n".join(texts))

                        slide_name = f"ppt/slides/_rels/slide{slide_num}.xml.rels"

                    notes_files = sorted(
                        [n for n in z.namelist() if n.startswith("ppt/notesSlides/")]
                    )

                    for notes_file in notes_files:
                        with z.open(notes_file) as f:
                            tree = ET.parse(f)

                        notes_text = []
                        for elem in tree.iter():
                            if elem.tag.endswith("}t") and elem.text and elem.text.strip():
                                notes_text.append(elem.text.strip())

                        if notes_text:
                            notes.append("\n".join(notes_text))

                    if "ppt/presentation.xml" in z.namelist():
                        with z.open("ppt/presentation.xml") as f:
                            pres_tree = ET.parse(f)

                        for elem in pres_tree.iter():
                            if elem.tag.endswith("}slide"):
                                layouts.append(
                                    elem.attrib.get(
                                        "{http://schemas.openxmlformats.org/presentationml/2006/main}name",
                                        "",
                                    )
                                )

            finally:
                os.unlink(tmp_path)

            output_parts = []

            for slide in slides:
                output_parts.append(slide)
                output_parts.append("")

            if notes:
                output_parts.append("## Speaker Notes")
                output_parts.append("")
                for i, note in enumerate(notes, 1):
                    output_parts.append(f"### Slide {i}")
                    output_parts.append(note)
                    output_parts.append("")

            return ParsedDocumentResult(
                text="\n---\n".join(output_parts),
                metadata={"slides": len(slides), "notes_slides": len(notes)},
            )
        except Exception as e:
            return ParsedDocumentResult(text="", error=str(e))

    def _parse_markdown(self, content: bytes) -> ParsedDocumentResult:
        text = content.decode("utf-8", errors="replace")
        links = re.findall(r"\[([^\]]+)\]\(([^)]+)\)", text)
        return ParsedDocumentResult(
            text=text,
            metadata={"link_count": len(links)},
            links=[l[1] for l in links],
        )

    def _parse_json(self, content: bytes) -> ParsedDocumentResult:
        import json

        try:
            data = json.loads(content.decode("utf-8"))
            formatted = json.dumps(data, indent=2, ensure_ascii=False)
            keys = list(data.keys()) if isinstance(data, dict) else None
            item_count = len(data) if isinstance(data, (list, dict)) else 0
            return ParsedDocumentResult(
                text=formatted,
                metadata={
                    "keys": keys,
                    "item_count": item_count,
                    "type": type(data).__name__,
                },
            )
        except json.JSONDecodeError as e:
            return ParsedDocumentResult(text="", error=f"JSON parse error: {e}")

    def _parse_yaml(self, content: bytes) -> ParsedDocumentResult:
        import yaml

        try:
            data = yaml.safe_load(content.decode("utf-8"))
            formatted = yaml.dump(data, default_flow_style=False, allow_unicode=True)
            keys = list(data.keys()) if isinstance(data, dict) else None
            item_count = len(data) if isinstance(data, (list, dict)) else 0
            return ParsedDocumentResult(
                text=formatted,
                metadata={
                    "keys": keys,
                    "item_count": item_count,
                    "type": type(data).__name__,
                },
            )
        except yaml.YAMLError as e:
            return ParsedDocumentResult(text="", error=f"YAML parse error: {e}")

    def _parse_xml(self, content: bytes) -> ParsedDocumentResult:
        import xml.etree.ElementTree as ET

        try:
            root = ET.fromstring(content.decode("utf-8"))
            text = self._xml_element_to_text(root)
            metadata = {
                "root_tag": root.tag,
                "child_count": len(root),
            }
            return ParsedDocumentResult(text=text, metadata=metadata)
        except ET.ParseError as e:
            return ParsedDocumentResult(text="", error=f"XML parse error: {e}")

    def _xml_element_to_text(self, element, indent: int = 0) -> str:
        parts = []
        prefix = "  " * indent
        if element.text and element.text.strip():
            parts.append(f"{prefix}{element.text.strip()}")
        for child in element:
            if child.tag:
                parts.append(f"{prefix}<{child.tag}>")
                parts.append(self._xml_element_to_text(child, indent + 1))
                parts.append(f"{prefix}</{child.tag}>")
        if element.tail and element.tail.strip():
            parts.append(f"{prefix}{element.tail.strip()}")
        return "\n".join(parts)

    def _parse_html(self, content: bytes) -> ParsedDocumentResult:
        from html.parser import HTMLParser

        class StructureExtractor(HTMLParser):
            def __init__(self):
                super().__init__()
                self.headings = []
                self.links = []
                self.list_items = []
                self.table_data = []
                self.current_table = []
                self.current_row = []
                self.in_list = False
                self.in_table = False
                self.in_td = False

            def handle_starttag(self, tag, attrs):
                attrs_dict = dict(attrs)
                if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
                    self.headings.append({"level": int(tag[1]), "text": ""})
                elif tag == "a" and "href" in attrs_dict:
                    self.links.append({"href": attrs_dict["href"], "text": ""})
                elif tag == "li":
                    self.in_list = True
                    self.list_items.append("")
                elif tag == "table":
                    self.in_table = True
                    self.current_table = []
                elif tag == "tr":
                    self.current_row = []
                elif tag in ("td", "th"):
                    self.in_td = True
                    self.current_row.append("")

            def handle_endtag(self, tag):
                if tag in ("h1", "h2", "h3", "h4", "h5", "h6") and self.headings:
                    self.headings[-1]["text"] = self.headings[-1]["text"].strip()
                elif tag == "a" and self.links and not self.links[-1]["text"]:
                    self.links[-1]["text"] = self.links[-1]["text"].strip()
                elif tag == "li":
                    self.in_list = False
                    if self.list_items:
                        self.list_items[-1] = self.list_items[-1].strip()
                elif tag == "table":
                    self.in_table = False
                    if self.current_table:
                        self.table_data.append(self.current_table)
                elif tag == "tr":
                    if self.current_row:
                        self.current_table.append(self.current_row)
                    self.current_row = []
                elif tag in ("td", "th"):
                    self.in_td = False

            def handle_data(self, data):
                data = data.strip()
                if not data:
                    return
                if self.headings:
                    self.headings[-1]["text"] += data + " "
                if self.links and not self.links[-1]["text"]:
                    self.links[-1]["text"] += data
                if self.in_list and self.list_items:
                    self.list_items[-1] += data + " "
                if self.in_td and self.current_row:
                    self.current_row[-1] += data

        try:
            html_text = content.decode("utf-8", errors="replace")
            extractor = StructureExtractor()
            extractor.feed(html_text)

            output_parts = []

            for h in extractor.headings:
                output_parts.append(f"{'#' * h['level']} {h['text']}")

            if extractor.headings:
                output_parts.append("")

            for item in extractor.list_items:
                if item:
                    output_parts.append(f"- {item}")

            if extractor.list_items:
                output_parts.append("")

            for table in extractor.table_data:
                for row in table:
                    output_parts.append("| " + " | ".join(row) + " |")
                output_parts.append("")

            for link in extractor.links[:10]:
                if link["text"]:
                    output_parts.append(f"[{link['text']}]({link['href']})")

            plain_text = re.sub(r"<[^>]+>", "", html_text)
            plain_text = re.sub(r"\n{3,}", "\n\n", plain_text)

            return ParsedDocumentResult(
                text=plain_text,
                metadata={
                    "headings": len(extractor.headings),
                    "links": len(extractor.links),
                    "list_items": len(extractor.list_items),
                    "tables": len(extractor.table_data),
                },
            )
        except Exception as e:
            return ParsedDocumentResult(text="", error=f"HTML parse error: {e}")


class HybridDocumentParser:
    """Hybrid parser: Docling → LlamaIndex → Fallback."""

    def __init__(
        self,
        prefer_llama_index: bool = True,
        use_docling: bool = True,
    ):
        self.llama = LlamaIndexParser()
        self.docling = DoclingParser()
        self.fallback = FallbackParser()
        self.prefer_llama_index = prefer_llama_index
        self.use_docling = use_docling

    async def parse(
        self,
        content: bytes,
        filename: str,
    ) -> ParsedDocumentResult:
        ext = os.path.splitext(filename)[1].lower()

        if ext == ".pdf" and self.use_docling and self.docling.available:
            result = await self.docling.parse(content)
            if result.text and not result.error:
                return result

        if self.prefer_llama_index and self.llama.available:
            result = await self.llama.parse(content, filename)
            if result.text and not result.error:
                return result

        return self.fallback.parse(content, filename)


# Global instance
_parser: Optional[HybridDocumentParser] = None


def get_document_parser() -> HybridDocumentParser:
    global _parser
    if _parser is None:
        _parser = HybridDocumentParser()
    return _parser


def is_llama_index_available() -> bool:
    return LLAMA_INDEX_AVAILABLE


def is_docling_available() -> bool:
    return DOCLING_AVAILABLE
